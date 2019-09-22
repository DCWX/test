# -*- utf-8 -*-
#@Time    :2019/9/422:37
#@Author  :无邪
#@File    :read_excel.py
#@Software:PyCharm
from openpyxl import load_workbook
from common.read_conf import ReadConf
from common.split_path import SplitPsth
class ReadExcel:
    """读取测试用例"""
    def __init__(self,casename,sheetname):
        """初始化文件名，表单名"""
        # print(SplitPsth().split_case())
        self.casename=casename
        self.sheet=sheetname
        self.wb=load_workbook(self.casename)
        self.sheet=self.wb[self.sheet]
    def choice(self):
        """判断配置文件中拿到的button值，选择读取的用例方式"""
        r = ReadConf(SplitPsth().split_conf())
        global s #声明全局变量
        s = r.read_choice()
        if r.read_choice() == "all":
            return self.read_excel()

        elif isinstance(eval(r.read_choice()), list):
            return self.read_single_case()
        else:
           print("你的选择有误，请重新选择")

    def read_excel(self):
        '''读取全部的用例'''
        case=[]
        for i in range(1, self.sheet.max_row+1):
            dicts={}
            dicts['Case_id']= self.sheet.cell(i+1, 1).value
            dicts['Module'] = self.sheet.cell(i+1, 2).value
            dicts['Title'] = self.sheet.cell(i+1, 3).value
            dicts['Method'] = self.sheet.cell(i+1, 4).value
            dicts['url'] = self.sheet.cell(i+1, 5).value
            dicts['Params'] = self.sheet.cell(i+1, 6).value
            dicts['ExpectedResult'] = self.sheet.cell(i+1, 7).value
            case.append(dicts)
            self.wb.close()
        return case
    def read_single_case(self):
        """根据配置文件读取用例"""
        case = []
        for i in eval(s):
            dicts = {}
            dicts['Case_id'] = self.sheet.cell(i + 1, 1).value
            dicts['Module'] = self.sheet.cell(i + 1, 2).value
            dicts['Title'] = self.sheet.cell(i + 1, 3).value
            dicts['Method'] = self.sheet.cell(i + 1, 4).value
            dicts['url'] = self.sheet.cell(i + 1, 5).value
            dicts['Params'] = self.sheet.cell(i + 1, 6).value
            dicts['ExpectedResult'] = self.sheet.cell(i + 1, 7).value
            case.append(dicts)
            self.wb.close()
        return case
    def write_excel(self,row,column,value):
        """写回测试结果"""
        self.wb.close()
        self.sheet.cell(row,column).value=value
        # self.wb.save(self.sheet)
        self.wb.save(self.casename)
        # self.wb.close()

if __name__ == '__main__':
    r=ReadExcel(SplitPsth().split_case(),"Sheet1")
    print(r.choice())
